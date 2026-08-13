#!/usr/bin/env python3
"""Build data.json from the live SharePoint Updated 4Miler Tracking.xlsx workbook.

Downloads the Marketing site workbook via Microsoft Graph (using an Azure CLI
user token), extracts the Leaderboard Tracker sheet, excludes partners, and
writes data.json for the static leaderboard.
"""
from __future__ import annotations

import argparse
import base64
import json
import subprocess
import sys
import tempfile
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data.json"

# SharePoint Doc.aspx link for Updated 4Miler Tracking.xlsx on the Marketing site.
SHAREPOINT_DOC_URL = (
    "https://saxllp.sharepoint.com/sites/Marketing/_layouts/15/Doc.aspx"
    "?sourcedoc={4462A57A-B085-4D46-B487-225D91621F37}"
    "&file=Updated%204Miler%20Tracking.xlsx"
)

TITLE = "INTERNAL COMPETITION"
SUBTITLE = "Current Standings · Fundraising & Participation Leaderboard"
DISCLAIMER = (
    "Total Participants and Total Funds Raised only count SAX employees "
    "in this competition below partner level."
)
PARTNERS_DISCLAIMER = (
    "Partners are ineligible to win any prizes and are therefore not "
    "reflected in the leaderboard."
)

# Activity point weights from the workbook headers / formulas.
WEIGHTS = {
    "raised": 1.0,
    "recruits": 25.0,
    "posts": 15.0,
    "registered": 25.0,
    "volunteered": 20.0,
    "shares": 10.0,
    "comments": 10.0,
}


def to_num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def get_graph_token() -> str:
    try:
        tok = subprocess.check_output(
            [
                "az",
                "account",
                "get-access-token",
                "--resource",
                "https://graph.microsoft.com",
                "--query",
                "accessToken",
                "-o",
                "tsv",
            ],
            text=True,
        ).strip()
    except subprocess.CalledProcessError as exc:
        raise SystemExit(
            "Failed to get Microsoft Graph token via Azure CLI. "
            "Run: az login --tenant a33e9b66-a6ef-43bf-9702-7cb4301d0a16"
        ) from exc
    if not tok:
        raise SystemExit("Azure CLI returned an empty Graph token. Run az login.")
    return tok


def graph_get_json(url: str, token: str) -> dict:
    req = urllib.request.Request(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(req) as resp:
        return json.load(resp)


def download_sharepoint_xlsx(dest: Path, token: str | None = None) -> dict:
    """Download workbook bytes and return driveItem metadata."""
    token = token or get_graph_token()
    share = "u!" + base64.urlsafe_b64encode(SHAREPOINT_DOC_URL.encode()).decode().rstrip("=")
    meta = graph_get_json(
        f"https://graph.microsoft.com/v1.0/shares/{share}/driveItem",
        token,
    )
    content_url = f"https://graph.microsoft.com/v1.0/shares/{share}/driveItem/content"
    req = urllib.request.Request(
        content_url,
        headers={"Authorization": f"Bearer {token}"},
    )
    # Follow redirects manually if needed; urlopen follows by default.
    with urllib.request.urlopen(req) as resp:
        data = resp.read()
    if len(data) < 1000 or data[:2] != b"PK":
        raise SystemExit(
            f"Downloaded file does not look like an xlsx (size={len(data)}). "
            f"First bytes: {data[:80]!r}"
        )
    dest.write_bytes(data)
    return meta


def compute_points(row: dict) -> float:
    """Recompute points from activity inputs (handles bonus-week 2x on posts)."""
    post_mult = to_num(row.get("post_multiplier")) or 1.0
    if post_mult <= 0:
        post_mult = 1.0
    return (
        WEIGHTS["raised"] * row["raised"]
        + WEIGHTS["recruits"] * row["recruits"]
        + WEIGHTS["posts"] * post_mult * row["posts"]
        + WEIGHTS["registered"] * row["registered"]
        + WEIGHTS["volunteered"] * row["volunteered"]
        + WEIGHTS["shares"] * row["shares"]
        + WEIGHTS["comments"] * row["comments"]
    )


def prize_for_rank(rank: int, points: float) -> str:
    if rank == 1:
        return "Gold"
    if rank <= 3:
        return "Silver"
    if rank <= 10:
        return "Bronze"
    if points >= 50:
        return "Participation"
    return ""


def dense_rank(people: list[dict], key) -> list[dict]:
    ranked: list[dict] = []
    prev = None
    rank = 0
    for idx, person in enumerate(people, 1):
        val = key(person)
        if prev is None or val != prev:
            rank = idx
            prev = val
        person = dict(person)
        person["rank"] = rank
        ranked.append(person)
    return ranked


def extract(path: Path) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Leaderboard Tracker" not in wb.sheetnames:
        raise SystemExit(
            f"Workbook missing 'Leaderboard Tracker' sheet. Found: {wb.sheetnames}"
        )
    ws = wb["Leaderboard Tracker"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise SystemExit("Leaderboard Tracker is empty")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    def col(*names: str) -> int | None:
        lowered = [h.lower() for h in header]
        for name in names:
            n = name.lower()
            for i, h in enumerate(lowered):
                if n == h or n in h:
                    return i
        return None

    idx_name = col("Employee Name", "Name") or 0
    idx_dept = col("Department") or 1
    idx_team = col("Team Name", "Team") or 2
    idx_raised = col("Total $ Raised", "Raised") or 3
    idx_recruits = col("Teammates Recruited", "Recruited") or 4
    idx_posts = col("Unique Social Posts", "Social Posts") or 5
    idx_registered = col("Registered") or 6
    idx_volunteered = col("Volunteered") or 7
    idx_shares = col("Share/Repost", "Share") or 8
    idx_comments = col("Comment/Follow", "Comment") or 9
    idx_points = col("Total Points", "Points") or 10
    idx_prize = col("Prize Tier", "Prize") or 12
    idx_mult = col("2x POINTS", "2x")
    idx_elig = col("Eligibility")

    people: list[dict] = []
    for row in rows[1:]:
        vals = list(row)
        if idx_name >= len(vals) or vals[idx_name] is None:
            continue
        name = str(vals[idx_name]).strip()
        if not name:
            continue

        elig = ""
        if idx_elig is not None and idx_elig < len(vals) and vals[idx_elig] is not None:
            elig = str(vals[idx_elig]).strip()
        if elig.lower() == "partner":
            continue

        dept = (
            str(vals[idx_dept]).strip()
            if idx_dept < len(vals) and vals[idx_dept]
            else "Unassigned"
        )
        team = (
            str(vals[idx_team]).strip()
            if idx_team < len(vals) and vals[idx_team]
            else ""
        )
        raised = to_num(vals[idx_raised] if idx_raised < len(vals) else 0)
        recruits = to_num(vals[idx_recruits] if idx_recruits < len(vals) else 0)
        posts = to_num(vals[idx_posts] if idx_posts < len(vals) else 0)
        registered = to_num(vals[idx_registered] if idx_registered < len(vals) else 0)
        volunteered = to_num(vals[idx_volunteered] if idx_volunteered < len(vals) else 0)
        shares = to_num(vals[idx_shares] if idx_shares < len(vals) else 0)
        comments = to_num(vals[idx_comments] if idx_comments < len(vals) else 0)
        sheet_points = to_num(vals[idx_points] if idx_points < len(vals) else 0)
        prize = (
            str(vals[idx_prize]).strip()
            if idx_prize < len(vals) and vals[idx_prize]
            else ""
        )
        post_multiplier = 1.0
        if idx_mult is not None and idx_mult < len(vals):
            post_multiplier = to_num(vals[idx_mult]) or 1.0

        if (
            raised <= 0
            and registered <= 0
            and sheet_points <= 0
            and (recruits + posts + volunteered + shares + comments) <= 0
        ):
            continue

        person = {
            "name": name,
            "department": dept or "Unassigned",
            "team": team,
            "raised": raised,
            "recruits": int(recruits),
            "posts": int(posts),
            "registered": int(registered),
            "volunteered": int(volunteered),
            "shares": int(shares),
            "comments": int(comments),
            "post_multiplier": post_multiplier,
            "sheet_points": sheet_points,
            "sheet_prize": prize,
        }
        # Prefer cached sheet points when present; otherwise compute.
        points = sheet_points if sheet_points > 0 else compute_points(person)
        person["points"] = points
        people.append(person)

    # Sort for stable output. Prefer workbook prize tiers (already formula-driven);
    # only recompute when the sheet prize cell is blank.
    people.sort(
        key=lambda p: (-to_num(p["points"]), -to_num(p["raised"]), p["name"].lower())
    )
    ranked = dense_rank(people, key=lambda p: to_num(p["points"]))
    out_people = []
    for p in ranked:
        prize = (p.get("sheet_prize") or "").strip()
        if not prize:
            prize = prize_for_rank(int(p["rank"]), to_num(p["points"]))
        out_people.append(
            {
                "name": p["name"],
                "department": p["department"],
                "team": p["team"],
                "raised": p["raised"],
                "recruits": p["recruits"],
                "posts": p["posts"],
                "registered": p["registered"],
                "volunteered": p["volunteered"],
                "shares": p["shares"],
                "comments": p["comments"],
                "points": p["points"],
                "prize": prize,
            }
        )

    return {
        "updatedAt": datetime.now(timezone.utc).isoformat(),
        "title": TITLE,
        "subtitle": SUBTITLE,
        "disclaimer": DISCLAIMER,
        "partnersDisclaimer": PARTNERS_DISCLAIMER,
        "source": {
            "file": "Updated 4Miler Tracking.xlsx",
            "sharePointUrl": SHAREPOINT_DOC_URL,
        },
        "people": out_people,
    }


def people_signature(data: dict) -> str:
    """Stable comparison payload ignoring updatedAt."""
    clone = {
        "title": data.get("title"),
        "subtitle": data.get("subtitle"),
        "disclaimer": data.get("disclaimer"),
        "partnersDisclaimer": data.get("partnersDisclaimer"),
        "people": data.get("people") or [],
    }
    return json.dumps(clone, sort_keys=True, separators=(",", ":"))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        help="Use a local xlsx instead of downloading from SharePoint",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=OUT,
        help=f"Output data.json path (default: {OUT})",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Always write output even when people data is unchanged",
    )
    args = parser.parse_args()

    meta = {}
    if args.xlsx:
        xlsx_path = args.xlsx
        if not xlsx_path.exists():
            raise SystemExit(f"Workbook not found: {xlsx_path}")
    else:
        with tempfile.TemporaryDirectory(prefix="4miler-sync-") as tmp:
            xlsx_path = Path(tmp) / "Updated 4Miler Tracking.xlsx"
            try:
                meta = download_sharepoint_xlsx(xlsx_path)
            except urllib.error.HTTPError as exc:
                body = exc.read().decode("utf-8", errors="replace")
                raise SystemExit(f"SharePoint download failed ({exc.code}): {body}") from exc
            data = extract(xlsx_path)
            # Keep going outside temp dir with data already extracted.
            return write_output(data, args.out, args.force, meta)

    data = extract(xlsx_path)
    return write_output(data, args.out, args.force, meta)


def write_output(data: dict, out: Path, force: bool, meta: dict) -> int:
    if meta:
        data.setdefault("source", {})
        data["source"]["itemId"] = meta.get("id")
        data["source"]["lastModifiedDateTime"] = meta.get("lastModifiedDateTime")
        data["source"]["webUrl"] = meta.get("webUrl")

    out.parent.mkdir(parents=True, exist_ok=True)
    previous = None
    if out.exists():
        try:
            previous = json.loads(out.read_text())
        except Exception:
            previous = None

    if (
        not force
        and previous is not None
        and people_signature(previous) == people_signature(data)
    ):
        # Preserve previous updatedAt when nothing meaningful changed.
        print(
            f"No leaderboard changes ({len(data['people'])} participants). "
            f"Left {out} unchanged."
        )
        return 0

    out.write_text(json.dumps(data, indent=2) + "\n")
    print(
        f"Wrote {out} ({len(data['people'])} participants, "
        f"updatedAt={data['updatedAt']})"
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        raise SystemExit(130)
